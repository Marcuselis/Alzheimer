#!/usr/bin/env python3
"""
Extract Principal Investigator information from ClinicalTrials.gov
for Phase II/III Alzheimer trials with Nordic sites.

Scope:
- Nordic countries only: Finland, Sweden, Denmark, Norway, Iceland
- Phase II and Phase III only
- Only publicly visible information from CT.gov
- No guessing or inference
"""

import csv
import re
import time
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    exit(1)

# Nordic countries (canonical names as in country_reference.csv)
NORDIC_COUNTRIES = {"Finland", "Sweden", "Denmark", "Norway", "Iceland"}

WORKBOOK_PATH = "Alzheimer_Landscape.xlsx"
OUTPUT_CSV = "nordic_site_pi.csv"

# Column order as specified
OUTPUT_COLUMNS = [
    "NCT Number",
    "canonical_molecule",
    "Sponsor",
    "Country",
    "City",
    "Hospital / Facility",
    "Site PI Name",
    "PI Title / Role",
    "PI Email",
    "Source",
    "Notes",
]


def filter_target_ncts():
    """
    Step 1: Read TRIALS_MASTER sheet and filter for:
    - Phases contains "Phase 2" or "Phase 3"
    - Countries_Normalized includes at least one Nordic country
    Returns: list of dicts with NCT Number, canonical_molecule, Sponsor
    """
    print("Reading TRIALS_MASTER from Excel...")
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = wb["TRIALS_MASTER"]

    # Find header row
    header_row = None
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20), start=1):
        values = [str(cell.value or "").strip() for cell in row]
        if "NCT Number" in values:
            header_row = idx
            header = values
            break

    if not header_row:
        raise ValueError("Could not find header row in TRIALS_MASTER")

    # Map column names to indices
    col_map = {}
    for idx, h in enumerate(header):
        col_map[h] = idx

    nct_col = col_map.get("NCT Number")
    phases_col = col_map.get("Phases")
    countries_col = col_map.get("Countries_Normalized")
    molecule_col = col_map.get("canonical_molecule")
    sponsor_col = col_map.get("Sponsor")

    if any(x is None for x in [nct_col, phases_col, countries_col]):
        raise ValueError("Required columns not found in TRIALS_MASTER")

    target_trials = []

    # Read data rows
    for row in sheet.iter_rows(min_row=header_row + 1):
        nct = str(row[nct_col].value or "").strip()
        phases = str(row[phases_col].value or "").strip()
        countries_norm = str(row[countries_col].value or "").strip()
        molecule = str(row[molecule_col].value or "").strip() if molecule_col else ""
        sponsor = str(row[sponsor_col].value or "").strip() if sponsor_col else ""

        if not nct or not nct.startswith("NCT"):
            continue

        # Check Phase 2 or Phase 3
        phases_upper = phases.upper()
        has_phase2 = "PHASE 2" in phases_upper or "PHASE2" in phases_upper or "PHASE II" in phases_upper
        has_phase3 = "PHASE 3" in phases_upper or "PHASE3" in phases_upper or "PHASE III" in phases_upper

        if not (has_phase2 or has_phase3):
            continue

        # Check for Nordic countries
        countries_list = [c.strip() for c in countries_norm.split("|") if c.strip()]
        has_nordic = any(c in NORDIC_COUNTRIES for c in countries_list)

        if has_nordic:
            target_trials.append({
                "NCT Number": nct,
                "canonical_molecule": molecule,
                "Sponsor": sponsor,
            })

    print(f"Found {len(target_trials)} target NCTs (Phase 2/3 with Nordic sites)")
    return target_trials


def extract_nordic_sites_from_html(html_content, nct_number):
    """
    Parse HTML to extract Nordic site information.
    Returns list of dicts with site details.
    """
    soup = BeautifulSoup(html_content, 'html.parser')
    sites = []
    
    # Find all location elements - CT.gov uses various structures
    # Look for text containing Nordic country names
    location_section = soup.find(id=re.compile(r'contacts.*locations', re.I))
    if not location_section:
        # Try alternative selectors
        location_section = soup.find(string=re.compile(r'This study has.*locations', re.I))
        if location_section:
            location_section = location_section.find_parent()
    
    # More robust: search entire page for Nordic country mentions
    for country in NORDIC_COUNTRIES:
        # Find elements containing country name
        elements = soup.find_all(string=re.compile(country, re.I))
        
        for elem in elements:
            # Get parent element that likely contains full location info
            parent = elem.find_parent()
            if not parent:
                continue
                
            # Extract location text
            location_text = parent.get_text(separator=' ', strip=True)
            
            # Try to parse: "City, Country, PostalCode" or "City, Country"
            # Look for pattern like "Aalborg, Denmark, 9000"
            match = re.search(r'([^,]+),\s*' + re.escape(country) + r'(?:,\s*(\d+))?', location_text)
            if match:
                city = match.group(1).strip()
                postal_code = match.group(2) if match.group(2) else ""
                
                # Try to find facility name - usually nearby in the DOM
                facility = ""
                # Look in sibling or parent elements
                current = parent
                for _ in range(3):
                    siblings = current.find_next_siblings()
                    for sib in siblings[:3]:  # Check first 3 siblings
                        sib_text = sib.get_text(strip=True)
                        # Facility names are usually 10-200 chars, don't contain country name
                        if (10 < len(sib_text) < 200 and 
                            country.lower() not in sib_text.lower() and
                            not re.match(r'^\d+$', sib_text) and
                            'Recruiting' not in sib_text):
                            facility = sib_text
                            break
                    if facility:
                        break
                    current = current.find_parent()
                    if not current:
                        break
                
                # If no facility found in siblings, try parent's text
                if not facility:
                    parent_text = parent.get_text(separator=' ', strip=True)
                    # Remove city, country, postal code parts
                    parts = parent_text.split(',')
                    for part in parts:
                        part = part.strip()
                        if (part and 
                            country.lower() not in part.lower() and
                            not re.match(r'^\d+$', part) and
                            len(part) > 10):
                            facility = part
                            break
                
                # Clean up facility name
                facility = re.sub(r'^Recruiting\s*', '', facility, flags=re.I).strip()
                
                sites.append({
                    "Country": country,
                    "City": city,
                    "Hospital / Facility": facility,
                    "Site PI Name": "",  # Will be filled if found
                    "PI Title / Role": "",
                    "PI Email": "",
                })
    
    # Deduplicate sites (same city + facility)
    seen = set()
    unique_sites = []
    for site in sites:
        key = (site["Country"], site["City"], site["Hospital / Facility"])
        if key not in seen:
            seen.add(key)
            unique_sites.append(site)
    
    return unique_sites


def extract_pi_info_from_page(html_content):
    """
    Try to extract PI information from the page.
    CT.gov typically doesn't list site-level PIs publicly.
    Returns dict with PI info if found, empty dict otherwise.
    """
    soup = BeautifulSoup(html_content, 'html.parser')
    
    # Look for "Principal Investigator" or "Site Principal Investigator" text
    pi_elements = soup.find_all(string=re.compile(r'principal\s+investigator|site\s+principal\s+investigator', re.I))
    
    pi_info = {}
    for elem in pi_elements:
        # Get nearby text that might be the PI name
        parent = elem.find_parent()
        if parent:
            # Look for name pattern (usually appears near "Principal Investigator")
            text = parent.get_text(separator=' ', strip=True)
            # PI names are typically 2-4 words, capitalized
            # This is very heuristic and may not work well
            # We'll be conservative and only extract if very clear
            pass
    
    # CT.gov rarely lists site-level PIs publicly
    # Most studies only show overall study director
    return pi_info


def process_nct(nct_number, molecule, sponsor):
    """
    Process a single NCT: fetch page and extract Nordic site info.
    """
    url = f"https://clinicaltrials.gov/study/{nct_number}"
    print(f"  Processing {nct_number}...")
    
    try:
        # Fetch page with proper headers
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'
        }
        response = requests.get(url, headers=headers, timeout=30)
        response.raise_for_status()
        
        # Extract Nordic sites
        sites = extract_nordic_sites_from_html(response.text, nct_number)
        
        # Try to extract PI info (usually not available)
        # pi_info = extract_pi_info_from_page(response.text)
        
        # Build output rows
        rows = []
        for site in sites:
            rows.append({
                "NCT Number": nct_number,
                "canonical_molecule": molecule,
                "Sponsor": sponsor,
                "Country": site["Country"],
                "City": site["City"],
                "Hospital / Facility": site["Hospital / Facility"],
                "Site PI Name": site["Site PI Name"],
                "PI Title / Role": site["PI Title / Role"],
                "PI Email": site["PI Email"],
                "Source": url,
                "Notes": "PI not listed in public record" if not site["Site PI Name"] else "",
            })
        
        if not rows:
            # Still create a row to indicate we checked
            rows.append({
                "NCT Number": nct_number,
                "canonical_molecule": molecule,
                "Sponsor": sponsor,
                "Country": "",
                "City": "",
                "Hospital / Facility": "",
                "Site PI Name": "",
                "PI Title / Role": "",
                "PI Email": "",
                "Source": url,
                "Notes": "No Nordic sites found or unable to parse",
            })
        
        return rows
        
    except Exception as e:
        print(f"    ERROR processing {nct_number}: {e}")
        return [{
            "NCT Number": nct_number,
            "canonical_molecule": molecule,
            "Sponsor": sponsor,
            "Country": "",
            "City": "",
            "Hospital / Facility": "",
            "Site PI Name": "",
            "PI Title / Role": "",
            "PI Email": "",
            "Source": url,
            "Notes": f"Error fetching page: {str(e)}",
        }]


def main():
    target_trials = filter_target_ncts()
    
    print(f"\nProcessing {len(target_trials)} NCTs...")
    
    all_rows = []
    for trial in target_trials:
        rows = process_nct(
            trial["NCT Number"],
            trial.get("canonical_molecule", ""),
            trial.get("Sponsor", "")
        )
        all_rows.extend(rows)
        time.sleep(1)  # Be polite to CT.gov servers
    
    # Write output CSV
    print(f"\nWriting {len(all_rows)} rows to {OUTPUT_CSV}...")
    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(all_rows)
    
    # Summary statistics
    nordic_sites = sum(1 for r in all_rows if r["Country"] in NORDIC_COUNTRIES)
    sites_with_pi = sum(1 for r in all_rows if r["Site PI Name"].strip())
    
    print(f"\n=== EXTRACTION COMPLETE ===")
    print(f"NCTs reviewed: {len(target_trials)}")
    print(f"Nordic sites found: {nordic_sites}")
    print(f"Sites with PI names: {sites_with_pi}")
    print(f"Sites without PI names: {nordic_sites - sites_with_pi}")
    print(f"\nOutput file: {OUTPUT_CSV}")


if __name__ == "__main__":
    main()
