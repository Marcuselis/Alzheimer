#!/usr/bin/env python3
"""
Extract Principal Investigator information using browser automation.
This script is designed to be called with browser MCP tools.
"""

import csv
import json
import re
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed")
    sys.exit(1)

NORDIC_COUNTRIES = {"Finland", "Sweden", "Denmark", "Norway", "Iceland"}
WORKBOOK_PATH = "Alzheimer_Landscape.xlsx"
OUTPUT_CSV = "nordic_site_pi.csv"

OUTPUT_COLUMNS = [
    "NCT Number",
    "canonical_molecule",
    "Sponsor",
    "Country",
    "City",
    "Hospital / Facility",
    "Site PI Name",
    "PI Title / Role",
    "PI Email",
    "Source",
    "Notes",
]


def filter_target_ncts():
    """Extract target NCTs from Excel file."""
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = wb["TRIALS_MASTER"]

    header_row = None
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20), start=1):
        values = [str(cell.value or "").strip() for cell in row]
        if "NCT Number" in values:
            header_row = idx
            header = values
            break

    if not header_row:
        raise ValueError("Could not find header row")

    col_map = {h: idx for idx, h in enumerate(header)}
    nct_col = col_map.get("NCT Number")
    phases_col = col_map.get("Phases")
    countries_col = col_map.get("Countries_Normalized")
    molecule_col = col_map.get("canonical_molecule")
    sponsor_col = col_map.get("Sponsor")

    target_trials = []
    for row in sheet.iter_rows(min_row=header_row + 1):
        nct = str(row[nct_col].value or "").strip()
        phases = str(row[phases_col].value or "").strip()
        countries_norm = str(row[countries_col].value or "").strip()
        molecule = str(row[molecule_col].value or "").strip() if molecule_col else ""
        sponsor = str(row[sponsor_col].value or "").strip() if sponsor_col else ""

        if not nct or not nct.startswith("NCT"):
            continue

        phases_upper = phases.upper()
        has_phase2 = "PHASE 2" in phases_upper or "PHASE2" in phases_upper
        has_phase3 = "PHASE 3" in phases_upper or "PHASE3" in phases_upper

        if not (has_phase2 or has_phase3):
            continue

        countries_list = [c.strip() for c in countries_norm.split("|") if c.strip()]
        has_nordic = any(c in NORDIC_COUNTRIES for c in countries_list)

        if has_nordic:
            target_trials.append({
                "NCT Number": nct,
                "canonical_molecule": molecule,
                "Sponsor": sponsor,
            })

    return target_trials


def extract_nordic_sites_from_browser_eval(result_json):
    """
    Parse the result from browser_evaluate JavaScript extraction.
    Returns list of site dicts.
    """
    try:
        data = json.loads(result_json) if isinstance(result_json, str) else result_json
        sites = []
        
        for site_data in data.get("nordicSites", []):
            country = site_data.get("country", "")
            city = site_data.get("city", "")
            facility = site_data.get("facility", "")
            
            # Clean up facility (remove "Recruiting" prefix if present)
            facility = re.sub(r'^Recruiting\s*', '', facility, flags=re.I).strip()
            
            if country in NORDIC_COUNTRIES and (city or facility):
                sites.append({
                    "Country": country,
                    "City": city,
                    "Hospital / Facility": facility,
                    "Site PI Name": "",
                    "PI Title / Role": "",
                    "PI Email": "",
                })
        
        return sites
    except Exception as e:
        print(f"Error parsing browser result: {e}")
        return []


if __name__ == "__main__":
    # This script outputs the target NCTs as JSON for the main extraction script
    trials = filter_target_ncts()
    print(json.dumps(trials, indent=2))
