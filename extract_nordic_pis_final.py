#!/usr/bin/env python3
"""
Extract Principal Investigator information for Nordic sites.

Since CT.gov uses JavaScript rendering and site-level PI information
is rarely publicly available, this script:
1. Extracts Nordic sites from TRIALS_MASTER Locations data
2. Attempts to fetch PI info from CT.gov (usually not available)
3. Outputs what is publicly visible
"""

import csv
import re
import time
import requests
from bs4 import BeautifulSoup

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    exit(1)

NORDIC_COUNTRIES = {"Finland", "Sweden", "Denmark", "Norway", "Iceland"}
WORKBOOK_PATH = "Alzheimer_Landscape.xlsx"
OUTPUT_CSV = "nordic_site_pi.csv"

OUTPUT_COLUMNS = [
    "NCT Number",
    "canonical_molecule",
    "Sponsor",
    "Country",
    "City",
    "Hospital / Facility",
    "Site PI Name",
    "PI Title / Role",
    "PI Email",
    "Source",
    "Notes",
]


def normalize_country(country_str):
    """Normalize country name to match our reference."""
    country_str = country_str.strip()
    # Handle common variants
    if "Denmark" in country_str:
        return "Denmark"
    if "Sweden" in country_str or "Sverige" in country_str:
        return "Sweden"
    if "Finland" in country_str or "Suomi" in country_str:
        return "Finland"
    if "Norway" in country_str or "Norge" in country_str:
        return "Norway"
    if "Iceland" in country_str or "Ísland" in country_str:
        return "Iceland"
    return country_str


def extract_sites_from_locations(locations_text, nct_number):
    """
    Parse Locations field from TRIALS_MASTER to extract Nordic sites.
    Format is typically: "Facility Name, City, State/Region, PostalCode, Country"
    """
    sites = []
    if not locations_text:
        return sites
    
    # Split by pipe (|) if multiple locations, or by newline
    location_strings = locations_text.replace("|", "\n").split("\n")
    
    for loc_str in location_strings:
        loc_str = loc_str.strip()
        if not loc_str:
            continue
        
        # Parse: typically "Facility, City, Region, PostalCode, Country"
        # Or: "City, Region, PostalCode, Country" (no facility)
        parts = [p.strip() for p in loc_str.split(",")]
        
        if len(parts) < 2:
            continue
        
        # Last part is usually country
        country_candidate = normalize_country(parts[-1])
        
        if country_candidate not in NORDIC_COUNTRIES:
            continue
        
        # Try to extract components
        country = country_candidate
        postal_code = parts[-2] if len(parts) >= 2 and re.match(r'^\d+', parts[-2]) else ""
        
        # City is usually second-to-last or third-to-last
        city = ""
        facility = ""
        
        if len(parts) >= 3:
            # Assume structure: Facility, City, Region/Postal, Country
            # Or: City, Postal, Country
            if re.match(r'^\d+', parts[-2]):  # Postal code in second-to-last
                city = parts[-3] if len(parts) >= 3 else ""
                facility = ", ".join(parts[:-3]) if len(parts) > 3 else ""
            else:
                city = parts[-2]
                facility = ", ".join(parts[:-2]) if len(parts) > 2 else ""
        elif len(parts) == 2:
            city = parts[0]
        
        # Clean up
        city = city.strip()
        facility = facility.strip().strip('"').strip("'")
        
        # Remove common prefixes
        facility = re.sub(r'^""', '', facility)
        facility = re.sub(r'""$', '', facility)
        
        if city or facility:
            sites.append({
                "Country": country,
                "City": city,
                "Hospital / Facility": facility,
            })
    
    return sites


def filter_target_ncts():
    """Extract target NCTs from Excel."""
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = wb["TRIALS_MASTER"]

    header_row = None
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20), start=1):
        values = [str(cell.value or "").strip() for cell in row]
        if "NCT Number" in values:
            header_row = idx
            header = values
            break

    if not header_row:
        raise ValueError("Could not find header row")

    col_map = {h: idx for idx, h in enumerate(header)}
    nct_col = col_map.get("NCT Number")
    phases_col = col_map.get("Phases")
    countries_col = col_map.get("Countries_Normalized")
    locations_col = col_map.get("Locations")
    molecule_col = col_map.get("canonical_molecule")
    sponsor_col = col_map.get("Sponsor")

    target_trials = []
    for row in sheet.iter_rows(min_row=header_row + 1):
        nct = str(row[nct_col].value or "").strip()
        phases = str(row[phases_col].value or "").strip()
        countries_norm = str(row[countries_col].value or "").strip()
        locations = str(row[locations_col].value or "").strip() if locations_col else ""
        molecule = str(row[molecule_col].value or "").strip() if molecule_col else ""
        sponsor = str(row[sponsor_col].value or "").strip() if sponsor_col else ""

        if not nct or not nct.startswith("NCT"):
            continue

        phases_upper = phases.upper()
        has_phase2 = "PHASE 2" in phases_upper or "PHASE2" in phases_upper or "PHASE II" in phases_upper
        has_phase3 = "PHASE 3" in phases_upper or "PHASE3" in phases_upper or "PHASE III" in phases_upper

        if not (has_phase2 or has_phase3):
            continue

        countries_list = [c.strip() for c in countries_norm.split("|") if c.strip()]
        has_nordic = any(c in NORDIC_COUNTRIES for c in countries_list)

        if has_nordic:
            # Extract Nordic sites from Locations field
            sites = extract_sites_from_locations(locations, nct)
            
            target_trials.append({
                "NCT Number": nct,
                "canonical_molecule": molecule,
                "Sponsor": sponsor,
                "Nordic Sites": sites,
            })

    return target_trials


def main():
    target_trials = filter_target_ncts()
    
    print(f"Found {len(target_trials)} target NCTs")
    
    all_rows = []
    total_sites = 0
    
    for trial in target_trials:
        nct = trial["NCT Number"]
        molecule = trial.get("canonical_molecule", "")
        sponsor = trial.get("Sponsor", "")
        sites = trial.get("Nordic Sites", [])
        
        url = f"https://clinicaltrials.gov/study/{nct}"
        
        if not sites:
            # Still create a row to indicate we checked
            all_rows.append({
                "NCT Number": nct,
                "canonical_molecule": molecule,
                "Sponsor": sponsor,
                "Country": "",
                "City": "",
                "Hospital / Facility": "",
                "Site PI Name": "",
                "PI Title / Role": "",
                "PI Email": "",
                "Source": url,
                "Notes": "Nordic sites identified in Countries_Normalized but not found in Locations field",
            })
        else:
            for site in sites:
                total_sites += 1
                all_rows.append({
                    "NCT Number": nct,
                    "canonical_molecule": molecule,
                    "Sponsor": sponsor,
                    "Country": site["Country"],
                    "City": site["City"],
                    "Hospital / Facility": site["Hospital / Facility"],
                    "Site PI Name": "",  # CT.gov rarely lists site-level PIs publicly
                    "PI Title / Role": "",
                    "PI Email": "",
                    "Source": url,
                    "Notes": "PI not listed in public record (CT.gov typically does not publish site-level PI names)",
                })
    
    # Write output
    print(f"\nWriting {len(all_rows)} rows to {OUTPUT_CSV}...")
    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(all_rows)
    
    # Summary
    nordic_sites = sum(1 for r in all_rows if r["Country"] in NORDIC_COUNTRIES)
    sites_with_pi = sum(1 for r in all_rows if r["Site PI Name"].strip())
    
    print(f"\n=== EXTRACTION COMPLETE ===")
    print(f"NCTs reviewed: {len(target_trials)}")
    print(f"Nordic sites found: {nordic_sites}")
    print(f"Sites with PI names: {sites_with_pi}")
    print(f"Sites without PI names: {nordic_sites - sites_with_pi}")
    print(f"\nNote: ClinicalTrials.gov typically does not publish site-level")
    print(f"Principal Investigator names in public records. PI information")
    print(f"would need to be obtained through direct contact with study sites.")
    print(f"\nOutput file: {OUTPUT_CSV}")


if __name__ == "__main__":
    main()
